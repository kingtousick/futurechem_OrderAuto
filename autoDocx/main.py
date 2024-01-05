import os
from flask import Flask, render_template, request, redirect
from tkinter import Tk, filedialog, Label, Button, Entry, Text
import sqlite3
import openpyxl
from docx import Document

# 나머지 코드

app = Flask(__name__)
DATABASE_NAME = "data.db"

def create_tables():
    # 데이터베이스 테이블 생성 함수
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()

    cursor.execute('''CREATE TABLE IF NOT EXISTS institutions (
                       seq INTEGER PRIMARY KEY AUTOINCREMENT,
                       institution_name TEXT,
                       institution_code TEXT
                       )''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS isotopes (
                       seq INTEGER PRIMARY KEY AUTOINCREMENT,
                       isotope_name TEXT,
                       quantity INTEGER
                       )''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS models (
                       seq INTEGER PRIMARY KEY AUTOINCREMENT,
                       model_name TEXT
                       )''')

    conn.commit()
    conn.close()

def insert_data(table_name, data):
    # 데이터베이스에 데이터 삽입 함수
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()

    if table_name == 'institutions':
        cursor.execute("INSERT INTO institutions (institution_name, institution_code) VALUES (?, ?)", data)
    elif table_name == 'isotopes':
        cursor.execute("INSERT INTO isotopes (isotope_name, quantity) VALUES (?, ?)", data)
    elif table_name == 'models':
        cursor.execute("INSERT INTO models (model_name) VALUES (?)", data)

    conn.commit()
    conn.close()

def read_data(table_name):
    # 데이터베이스에서 데이터 조회 함수
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()

    cursor.execute(f"SELECT * FROM {table_name}")
    data = cursor.fetchall()

    conn.close()
    return data

def modify_word_file(institution_name, institution_code, isotope_name, quantity, model_name, manpower, purchase_date):
    # Word 파일 수정 및 저장 함수
    document = Document("Purchase_Request_Template.docx")

    for paragraph in document.paragraphs:
        if "Institution" in paragraph.text:
            paragraph.text = f"Institution: {institution_name} - {institution_code}"
        elif "Isotope" in paragraph.text:
            paragraph.text = f"Isotope: {isotope_name} - Quantity: {quantity}"
        elif "Model" in paragraph.text:
            paragraph.text = f"Model: {model_name}"
        elif "Manpower" in paragraph.text:
            paragraph.text = f"Manpower: {manpower}"
        elif "Purchase Date" in paragraph.text:
            paragraph.text = f"Purchase Date: {purchase_date}"

    new_file_name = f"Purchase_Request_{institution_name}_{isotope_name}_{model_name}.docx"
    document.save(new_file_name)

def main():
    create_tables()

    root = Tk()
    root.title("Purchase Request Manager")

    def upload_excel():
        # 엑셀 파일 업로드 함수
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = openpyxl.load_workbook(file_path)

            institutions_sheet = wb['Institutions']
            isotopes_sheet = wb['Isotopes']
            models_sheet = wb['Models']

            for row in institutions_sheet.iter_rows(min_row=2, values_only=True):
                insert_data('institutions', row)

            for row in isotopes_sheet.iter_rows(min_row=2, values_only=True):
                insert_data('isotopes', row)

            for row in models_sheet.iter_rows(min_row=2, values_only=True):
                insert_data('models', row)

            print("Data uploaded successfully!")

    def modify_word():
        # Word 파일 수정 함수
        institution_name = entry_institution_name.get()
        institution_code = entry_institution_code.get()
        isotope_name = entry_isotope_name.get()
        quantity = entry_quantity.get()
        model_name = entry_model_name.get()
        manpower = entry_manpower.get()
        purchase_date = entry_purchase_date.get()

        modify_word_file(institution_name, institution_code, isotope_name, quantity, model_name, manpower, purchase_date)
        print(f"Word document modified and saved as: Purchase_Request_{institution_name}_{isotope_name}_{model_name}.docx")

    Label(root, text="Institution Name:").grid(row=0, column=0)
    Label(root, text="Institution Code:").grid(row=1, column=0)
    Label(root, text="Isotope Name:").grid(row=2, column=0)
    Label(root, text="Quantity:").grid(row=3, column=0)
    Label(root, text="Model Name:").grid(row=4, column=0)
    Label(root, text="Manpower:").grid(row=5, column=0)
    Label(root, text="Purchase Date:").grid(row=6, column=0)

    entry_institution_name = Entry(root)
    entry_institution_code = Entry(root)
    entry_isotope_name = Entry(root)
    entry_quantity = Entry(root)
    entry_model_name = Entry(root)
    entry_manpower = Entry(root)
    entry_purchase_date = Entry(root)

    entry_institution_name.grid(row=0, column=1)
    entry_institution_code.grid(row=1, column=1)
    entry_isotope_name.grid(row=2, column=1)
    entry_quantity.grid(row=3, column=1)
    entry_model_name.grid(row=4, column=1)
    entry_manpower.grid(row=5, column=1)
    entry_purchase_date.grid(row=6, column=1)

    Button(root, text="Upload Excel", command=upload_excel).grid(row=7, column=0, columnspan=2)
    Button(root, text="Modify Word Document", command=modify_word).grid(row=8, column=0, columnspan=2)

    root.mainloop()

if __name__ == '__main__':
    main()
